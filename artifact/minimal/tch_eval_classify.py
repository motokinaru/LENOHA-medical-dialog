#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tch_eval_classify.py
Publication-grade metrics for the LENOHA classifier.
"""

import os, json, argparse, math, random, sys, subprocess
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer
from sklearn.metrics import (
    roc_curve, auc, confusion_matrix, precision_recall_fscore_support,
    accuracy_score, average_precision_score, precision_recall_curve
)

# ---- utils ----
def set_seeds(seed: int):
    if seed is None: 
        return
    random.seed(seed)
    np.random.seed(seed)
    try:
        import torch
        torch.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.use_deterministic_algorithms(False)  # STのencodeは完全決定論でなくてOK
    except Exception:
        pass

def wilson_ci(k, n, z=1.96):
    if n == 0:
        return (0.0, 0.0, 0.0)
    phat = k/n
    denom = 1 + z*z/n
    centre = phat + z*z/(2*n)
    adj = z * np.sqrt((phat*(1-phat) + z*z/(4*n))/n)
    lower = (centre - adj)/denom
    upper = (centre + adj)/denom
    return (phat, max(0.0, float(lower)), min(1.0, float(upper)))

def load_faq_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    qq, aa = [], []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1]:
            qq.append(str(row[0]).strip()); aa.append(str(row[1]).strip())
    if not qq:
        raise ValueError("No FAQs found in Excel (col1=Q, col2=A).")
    return qq, aa

def build_embeddings(model, model_name, faq_q, batch_size=32):
    is_e5 = "e5" in model_name.lower()
    passages = [f"passage: {q}" if is_e5 else q for q in faq_q]
    emb = model.encode(
        passages, normalize_embeddings=True, convert_to_numpy=True,
        show_progress_bar=True, batch_size=batch_size
    )
    return emb

def score_and_predict(model, model_name, faq_emb, inputs, threshold, batch_size=32):
    is_e5 = "e5" in model_name.lower()
    # バッチでクエリ埋め込み
    queries = [f"query: {t}" if is_e5 else t for t in inputs]
    q_emb = model.encode(
        queries, normalize_embeddings=True, convert_to_numpy=True,
        show_progress_bar=True, batch_size=batch_size
    )
    # コサイン類似（正規化済みなので内積）
    scores = (q_emb @ faq_emb.T).max(axis=1).astype(np.float32)
    preds = (scores >= float(threshold)).astype(int)
    return scores, preds

def mcnemar_test(y_true, pred_a, pred_b):
    y_true = np.asarray(y_true).astype(int)
    pred_a = np.asarray(pred_a).astype(int)
    pred_b = np.asarray(pred_b).astype(int)
    if y_true.shape[0] != pred_a.shape[0] or y_true.shape[0] != pred_b.shape[0]:
        raise ValueError("compare_preds length mismatch with ground truth")
    a_correct = (pred_a == y_true)
    b_correct = (pred_b == y_true)
    b01 = int(np.sum((a_correct == True) & (b_correct == False)))
    c10 = int(np.sum((a_correct == False) & (b_correct == True)))
    n = b01 + c10
    # 継続性補正つきのMcNemar
    if n == 0:
        chi2 = 0.0
        p = 1.0
    else:
        chi2 = (abs(b01 - c10) - 1)**2 / n
        # df=1 の厳密な生存関数: p = erfc(sqrt(chi2/2))
        p = float(math.erfc(math.sqrt(max(chi2, 0.0) / 2.0)))
    p_a = float(np.mean(a_correct))
    p_b = float(np.mean(b_correct))
    h = 2.0 * (math.asin(math.sqrt(p_a)) - math.asin(math.sqrt(p_b)))
    return {"b01": b01, "c10": c10, "chi2_cc": chi2, "p": p, "cohens_h": h, "acc_a": p_a, "acc_b": p_b}

def log_env(out_dir):
    try:
        info = {}
        for mod in ["numpy","pandas","scikit_learn","sentence_transformers"]:
            try:
                m = __import__(mod if mod!="scikit_learn" else "sklearn")
                v = getattr(m, "__version__", "unknown")
                info[mod] = v
            except Exception:
                info[mod] = "unavailable"
        with open(os.path.join(out_dir, "env.json"), "w", encoding="utf-8") as f:
            json.dump(info, f, ensure_ascii=False, indent=2)
        # requirements ロック
        try:
            req = subprocess.check_output([sys.executable, "-m", "pip", "freeze"], text=True)
            with open(os.path.join(out_dir, "requirements.lock"), "w", encoding="utf-8") as f:
                f.write(req)
        except Exception:
            pass
    except Exception:
        pass

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--faq_xlsx", required=True)
    ap.add_argument("--input_csv", required=True)
    ap.add_argument("--st_model", default="intfloat/multilingual-e5-large-instruct")
    ap.add_argument("--threshold", type=float, default=0.905)
    ap.add_argument("--out_dir", default="eval_out")
    ap.add_argument("--compare_preds", default="")
    ap.add_argument("--batch_size", type=int, default=64)
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--log_env", action="store_true")
    args = ap.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)
    set_seeds(args.seed)

    df = pd.read_csv(args.input_csv, encoding="utf-8-sig")
    if not {"input","label"}.issubset(df.columns):
        raise SystemExit("input_csv must contain 'input' and 'label' columns")
    texts = df["input"].astype(str).tolist()
    y = df["label"].astype(int).to_numpy()

    st = SentenceTransformer(args.st_model)
    faq_q, _ = load_faq_xlsx(args.faq_xlsx)
    faq_emb = build_embeddings(st, args.st_model, faq_q, batch_size=args.batch_size)

    scores, preds = score_and_predict(st, args.st_model, faq_emb, texts, args.threshold, batch_size=args.batch_size)

    tn, fp, fn, tp = confusion_matrix(y, preds, labels=[0,1]).ravel()
    acc = accuracy_score(y, preds)
    prec, rec, f1, _ = precision_recall_fscore_support(y, preds, average="binary", zero_division=0)
    spec = tn / (tn + fp) if (tn+fp) else 0.0
    bal_acc = (rec + spec) / 2.0

    acc_ci = wilson_ci(int(np.sum(y==preds)), len(y))
    prec_ci = wilson_ci(int(tp), int(tp+fp)) if (tp+fp)>0 else (0.0,0.0,0.0)
    rec_ci  = wilson_ci(int(tp), int(tp+fn)) if (tp+fn)>0 else (0.0,0.0,0.0)
    spec_ci = wilson_ci(int(tn), int(tn+fp)) if (tn+fp)>0 else (0.0,0.0,0.0)

    # ROC/AUC（単一クラスのガード）
    if len(np.unique(y)) > 1:
        fpr, tpr, thr = roc_curve(y, scores)
        auc_val = auc(fpr, tpr)
        roc_df = pd.DataFrame({"fpr":fpr, "tpr":tpr, "thresholds":thr})
        roc_df.to_csv(os.path.join(args.out_dir,"roc_points.csv"), index=False, encoding="utf-8-sig")
    else:
        auc_val = float("nan")

    # PR/AUPRC
    try:
        prec_points, rec_points, pr_thr = precision_recall_curve(y, scores)
        auprc = average_precision_score(y, scores)
        pr_df = pd.DataFrame({"precision":prec_points, "recall":rec_points})
        pr_df.to_csv(os.path.join(args.out_dir,"pr_points.csv"), index=False, encoding="utf-8-sig")
    except Exception:
        auprc = float("nan")

    out_preds = pd.DataFrame({"input":texts, "label":y, "score":scores, "pred":preds})
    out_preds.to_csv(os.path.join(args.out_dir,"preds.csv"), index=False, encoding="utf-8-sig")

    types = np.where((preds==1)&(y==0), "FP", np.where((preds==0)&(y==1), "FN",""))
    mis = out_preds[types!=""].copy()
    mis["type"] = types[types!=""]
    mis.to_csv(os.path.join(args.out_dir,"misclassified.csv"), index=False, encoding="utf-8-sig")

    metrics = {
        "n": int(len(y)),
        "threshold": float(args.threshold),
        "accuracy": {"point": float(acc_ci[0]), "ci95": [float(acc_ci[1]), float(acc_ci[2]) ]},
        "precision": {"point": float(prec_ci[0]), "ci95": [float(prec_ci[1]), float(prec_ci[2])]},
        "recall": {"point": float(rec_ci[0]), "ci95": [float(rec_ci[1]), float(rec_ci[2])]},
        "specificity": {"point": float(spec_ci[0]), "ci95": [float(spec_ci[1]), float(spec_ci[2])]},
        "f1": {"point": float(f1), "ci95": [None, None]},
        "balanced_accuracy": float(bal_acc),
        "auc": float(auc_val),
        "auprc": float(auprc),
        "confusion": {"tn": int(tn), "fp": int(fp), "fn": int(fn), "tp": int(tp)}
    }
    with open(os.path.join(args.out_dir,"metrics.json"), "w", encoding="utf-8") as f:
        json.dump(metrics, f, ensure_ascii=False, indent=2)

    flat = {
        "n": len(y), "threshold": args.threshold, "accuracy": acc_ci[0],
        "acc_ci_lo": acc_ci[1], "acc_ci_hi": acc_ci[2],
        "precision": prec_ci[0], "prec_ci_lo": prec_ci[1], "prec_ci_hi": prec_ci[2],
        "recall": rec_ci[0], "rec_ci_lo": rec_ci[1], "rec_ci_hi": rec_ci[2],
        "specificity": spec_ci[0], "spec_ci_lo": spec_ci[1], "spec_ci_hi": spec_ci[2],
        "f1": f1, "balanced_accuracy": bal_acc, "auc": auc_val, "auprc": auprc,
        "tn": tn, "fp": fp, "fn": fn, "tp": tp
    }
    pd.DataFrame([flat]).to_csv(os.path.join(args.out_dir,"metrics.csv"), index=False, encoding="utf-8-sig")

    if args.compare_preds:
        b = pd.read_csv(args.compare_preds, encoding="utf-8-sig")
        if "pred" not in b.columns:
            raise SystemExit("compare_preds must have column 'pred'")
        if len(b) != len(y):
            raise SystemExit("compare_preds must have the same number of rows as input_csv")
        res = mcnemar_test(y, preds, b["pred"].astype(int).to_numpy())
        with open(os.path.join(args.out_dir,"mcnemar.json"), "w", encoding="utf-8") as f:
            json.dump(res, f, ensure_ascii=False, indent=2)

    if args.log_env:
        log_env(args.out_dir)

    print("Saved:", os.path.abspath(args.out_dir))

if __name__ == "__main__":
    main()

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import logging
from logging.handlers import RotatingFileHandler
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
import os
import argparse
from typing import List, Dict

# Configuration
CONFIG_FILE = "config.json"

# Load configuration from file
import json
with open(CONFIG_FILE, "r") as f:
    config = json.load(f)

# Define model definitions and domain configurations
MODEL_DEFINITIONS = config["model_definitions"]
DOMAIN_CONFIGS = config["domain_configs"]

# Set up logger
def setup_logger(name: str, log_file: str, level: int = logging.INFO) -> logging.Logger:
    """
    Set up a logger with a rotating file handler.

    Args:
        name (str): Logger name.
        log_file (str): Path to log file.
        level (int, optional): Logging level. Defaults to logging.INFO.

    Returns:
        logging.Logger: Logger instance.
    """
    logger = logging.getLogger(name)
    logger.setLevel(level)
    if logger.hasHandlers():
        logger.handlers.clear()
    
    fh = RotatingFileHandler(log_file, maxBytes=1000000, backupCount=3, encoding='utf-8')
    fh.setLevel(level)
    ch = logging.StreamHandler()
    ch.setLevel(level)
    
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger

# Define functions
def load_st_models(model_defs: Dict[str, str], logger: logging.Logger) -> Dict[str, SentenceTransformer]:
    """
    Load Sentence Transformer models.

    Args:
        model_defs (Dict[str, str]): Model definitions.
        logger (logging.Logger): Logger instance.

    Returns:
        Dict[str, SentenceTransformer]: Loaded models.
    """
    logger.info("Loading Sentence Transformer models...")
    loaded_models = {}
    for name, model_name_hf in model_defs.items():
        try:
            logger.info(f"  Loading {name} ({model_name_hf})...")
            loaded_models[name] = SentenceTransformer(model_name_hf)
        except Exception as e_model_load:
            logger.error(f"Failed to load model {name} ({model_name_hf}): {e_model_load}")
            raise
    logger.info("All ST models loaded successfully.")
    return loaded_models

def load_faq_data(faq_fp: str, logger: logging.Logger) -> List[str]:
    """
    Load FAQ data from an Excel file.

    Args:
        faq_fp (str): Path to FAQ Excel file.
        logger (logging.Logger): Logger instance.

    Returns:
        List[str]: List of FAQ questions.
    """
    logger.info(f"Loading FAQ data from: {faq_fp}")
    faq_questions_list = []
    try:
        if not os.path.exists(faq_fp):
            logger.error(f"FAQ file not found: {faq_fp}")
            return None
        workbook = load_workbook(filename=faq_fp, read_only=True)
        sheet = workbook.active
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if row[0] is not None and str(row[0]).strip() != "":
                faq_questions_list.append(str(row[0]).strip())
        if not faq_questions_list:
            logger.error(f"No questions found in FAQ file: {faq_fp}. Ensure the first column contains questions starting from the second row.")
            return None
        logger.info(f"Loaded {len(faq_questions_list)} FAQ questions.")
        return faq_questions_list
    except Exception as e_faq_load:
        logger.error(f"Failed to load FAQ data from {faq_fp}: {e_faq_load}")
        return None

def embed_faqs(faq_questions_list: List[str], st_models_dict: Dict[str, SentenceTransformer], logger: logging.Logger) -> Dict[str, np.ndarray]:
    """
    Embed FAQ questions using Sentence Transformer models.

    Args:
        faq_questions_list (List[str]): List of FAQ questions.
        st_models_dict (Dict[str, SentenceTransformer]): Loaded models.
        logger (logging.Logger): Logger instance.

    Returns:
        Dict[str, np.ndarray]: Embedded FAQ questions.
    """
    logger.info("Embedding FAQ questions for all models...")
    faq_embeddings_dict = {}
    for name, model in st_models_dict.items():
        logger.info(f"  Embedding FAQs for {name}...")
        try:
            faq_embeddings_dict[name] = model.encode(faq_questions_list, show_progress_bar=True)
        except Exception as e_embed:
            logger.error(f"Failed to embed FAQs for model {name}: {e_embed}")
            raise
    logger.info("FAQ embedding complete.")
    return faq_embeddings_dict

def classify_utterance(user_input_text: str, st_models_dict: Dict[str, SentenceTransformer], faq_embeddings_dict: Dict[str, np.ndarray], model_thresholds_dict: Dict[str, float], logger: logging.Logger) -> Dict[str, Dict[str, float]]:
    """
    Classify an utterance using Sentence Transformer models.

    Args:
        user_input_text (str): User input text.
        st_models_dict (Dict[str, SentenceTransformer]): Loaded models.
        faq_embeddings_dict (Dict[str, np.ndarray]): Embedded FAQ questions.
        model_thresholds_dict (Dict[str, float]): Model thresholds.
        logger (logging.Logger): Logger instance.

    Returns:
        Dict[str, Dict[str, float]]: Classification results.
    """
    classification_results = {}
    for model_name, model_instance in st_models_dict.items():
        pred_label = 0  # Default to casual
        max_score = 0.0 # Default score
        status_message = "Casual Conversation (Default)"

        try:
            if model_name not in faq_embeddings_dict or faq_embeddings_dict[model_name].shape[0] == 0:
                status_message = "Error (FAQ empty)"
                logger.warning(f"No FAQ embeddings for model {model_name}. Input: '{user_input_text}' classified as casual by default.")
            else:
                user_vec = model_instance.encode([user_input_text])
                sims = cosine_similarity(user_vec, faq_embeddings_dict[model_name])[0]
                
                if len(sims) > 0:
                    max_score = float(np.max(sims))
                    model_specific_threshold = model_thresholds_dict.get(model_name)
                    
                    if model_specific_threshold is None:
                        status_message = "Error (No threshold)"
                        logger.error(f"Optimal threshold not found for model {model_name}. Cannot classify input '{user_input_text}'.")
                        pred_label = -1 # Indicate error
                    elif max_score >= model_specific_threshold:
                        pred_label = 1
                        status_message = "Clinical Question (FAQ Match)"
                    else:
                        pred_label = 0
                        status_message = "Casual Conversation"
                else:
                    status_message = "Error (Similarity calculation failed)"
                    logger.warning(f"Could not calculate similarities for model {model_name} with input '{user_input_text}'. Classified as casual by default.")
        except Exception as e_classify:
            status_message = "Error (Exception during classification)"
            logger.error(f"Error during classification for model {model_name} with input '{user_input_text}': {e_classify}")
            pred_label = -1
            max_score = 0.0
        classification_results[model_name] = {"label": pred_label, "score": max_score, "status": status_message}
    return classification_results

def load_test_data(input_csv_fp: str, logger: logging.Logger, num_samples_to_load: int = None) -> Tuple[List[str], List[int]]:
    """
    Load test data from a CSV file.

    Args:
        input_csv_fp (str): Path to input CSV file.
        logger (logging.Logger): Logger instance.
        num_samples_to_load (int, optional): Number of samples to load. Defaults to None.

    Returns:
        Tuple[List[str], List[int]]: Loaded test data.
    """
    logger.info(f"Loading test data from: {input_csv_fp}")
    try:
        if not os.path.exists(input_csv_fp):
            logger.error(f"Test data CSV file not found: {input_csv_fp}")
            return None, None
        input_df = pd.read_csv(input_csv_fp, encoding="utf-8-sig")
        if "input" not in input_df.columns or "label" not in input_df.columns:
            logger.error(f"Input CSV '{input_csv_fp}' must contain 'input' and 'label' columns. Found: {input_df.columns.tolist()}")
            return None, None
        if num_samples_to_load is None:
            num_samples_to_load = len(input_df)
        inputs = input_df["input"].astype(str).tolist()[:num_samples_to_load]
        ground_truth_labels = input_df["label"].astype(int).tolist()[:num_samples_to_load]
        logger.info(f"Loaded {len(inputs)} utterances for evaluation.")
        return inputs, ground_truth_labels
    except Exception as e_load_test:
        logger.error(f"Failed to load test data from {input_csv_fp}: {e_load_test}")
        return None, None

def main():
    parser = argparse.ArgumentParser(description="Sentence Transformer Model Evaluation Script for Medical Dialogue Classification")
    parser.add_argument("--domain", type=str, required=True, choices=list(DOMAIN_CONFIGS.keys()),
                        help=f"Target domain for evaluation: {', '.join(DOMAIN_CONFIGS.keys())}")
    parser.add_argument("--faq_file", type=str, help="Path to the FAQ Excel file (overrides default for the selected domain). Assumed to be in the script's directory if only filename is given.")
    parser.add_argument("--input_csv", type=str, help="Path to the input CSV (test data) file (overrides default for the selected domain). Assumed to be in the script's directory if only filename is given.")
    parser.add_argument("--max_samples", type=int, default=None, help="Maximum number of samples to process from the input CSV. Default is all.")
    parser.add_argument("--output_dir", type=str, default=".", help="Directory to save the output CSV file. Default is current directory.")

    cli_args = parser.parse_args()
    TARGET_DOMAIN = cli_args.domain

    # Initialize logger
    logger = setup_logger(f"{TARGET_DOMAIN.upper()}_Eval", DOMAIN_CONFIGS[TARGET_DOMAIN]['log_file'])

    # Load FAQ data
    faq_file_path = cli_args.faq_file if cli_args.faq_file else os.path.join(os.getcwd(), DOMAIN_CONFIGS[TARGET_DOMAIN]['faq_file'])
    faq_questions = load_faq_data(faq_file_path, logger)

    # Load test data
    input_csv_path = cli_args.input_csv if cli_args.input_csv else os.path.join(os.getcwd(), DOMAIN_CONFIGS[TARGET_DOMAIN]['input_csv'])
    inputs_for_df, ground_truth_labels_for_df = load_test_data(input_csv_path, logger, num_samples_to_load=cli_args.max_samples)

    # Load models
    st_models = load_st_models(MODEL_DEFINITIONS, logger)

    # Embed FAQs
    faq_embeddings = embed_faqs(faq_questions, st_models, logger)

    # Classify utterances
    all_model_predictions = {}
    for i, user_input in enumerate(inputs_for_df):
        gt_label = ground_truth_labels_for_df[i]
        classification_outputs = classify_utterance(user_input, st_models, faq_embeddings, DOMAIN_CONFIGS[TARGET_DOMAIN]['thresholds'], logger)
        all_model_predictions[user_input] = classification_outputs

    # Save output
    output_data_dict = {"Input": inputs_for_df, "GroundTruth": ground_truth_labels_for_df}
    for model_key in st_models.keys():
        output_data_dict[f"{model_key}_PredLabel"] = [all_model_predictions[input_][model_key]["label"] for input_ in inputs_for_df]
        output_data_dict[f"{model_key}_Score"] = [all_model_predictions[input_][model_key]["score"] for input_ in inputs_for_df]
    
    output_df_final = pd.DataFrame(output_data_dict)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_output_filename_only = f"{DOMAIN_CONFIGS[TARGET_DOMAIN]['output_prefix']}_{timestamp}.csv"
    final_output_full_path = os.path.join(cli_args.output_dir, final_output_filename_only)
    
    try:
        output_df_final.to_csv(final_output_full_path, index=False, encoding="utf-8-sig")
        logger.info(f"Classification results saved to CSV: {final_output_full_path}")
        print(f"\n Classification results (with scores and predicted labels) saved to CSV → {final_output_full_path}")
    except Exception as e_save_csv:
        logger.error(f"Failed to save output CSV to {final_output_full_path}: {e_save_csv}")
        print(f"ERROR: Failed to save output CSV to {final_output_full_path}: {e_save_csv}")

if __name__ == "__main__":
    main()
